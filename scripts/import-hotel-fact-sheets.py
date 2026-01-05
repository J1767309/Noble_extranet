#!/usr/bin/env python3
"""
Import Hotel Fact Sheets from Excel to Supabase

This script parses the Excel file containing hotel fact sheets and generates
SQL INSERT statements to populate the hotel_fact_sheets table.

Usage:
    python3 scripts/import-hotel-fact-sheets.py
"""

import openpyxl
import json
import re
import os

# Path to the Excel file
EXCEL_PATH = "Hotel Fact Sheet/All Hotel Fact Sheets 5.2025.xlsx"
OUTPUT_SQL_PATH = "supabase/migrations/029_import_hotel_fact_sheets.sql"

def clean_value(val):
    """Clean a cell value - strip whitespace and handle None"""
    if val is None:
        return None
    if isinstance(val, str):
        val = val.strip()
        # Remove leading indentation
        val = re.sub(r'^\s+', '', val)
        return val if val else None
    return val

def extract_number(text):
    """Extract a number from text like '1969' or '104 surface spaces'"""
    if text is None:
        return None
    if isinstance(text, (int, float)):
        return int(text)
    match = re.search(r'(\d+)', str(text))
    return int(match.group(1)) if match else None

def extract_year(text):
    """Extract a 4-digit year from text"""
    if text is None:
        return None
    if isinstance(text, (int, float)):
        return int(text) if 1900 < int(text) < 2100 else None
    match = re.search(r'(19|20)\d{2}', str(text))
    return int(match.group(0)) if match else None

def parse_city_state_zip(text):
    """Parse 'City, ST XXXXX' format"""
    if not text:
        return None, None, None
    text = clean_value(text)
    if not text:
        return None, None, None

    # Pattern: City, State ZIP
    match = re.match(r'(.+?),\s*([A-Z]{2})\s*(\d{5}(?:-\d{4})?)?', text)
    if match:
        return match.group(1).strip(), match.group(2), match.group(3)
    return text, None, None

def parse_county_submarket(text):
    """Parse 'County/Submarket' format"""
    if not text:
        return None, None
    text = clean_value(text)
    if not text:
        return None, None

    if '/' in text:
        parts = text.split('/', 1)
        return parts[0].strip(), parts[1].strip()
    return text, None

def escape_sql(val, is_jsonb=False):
    """Escape a value for SQL"""
    if val is None:
        return "NULL"
    if isinstance(val, bool):
        return "TRUE" if val else "FALSE"
    if isinstance(val, (int, float)):
        return str(val)
    if isinstance(val, list):
        if not val:
            return "NULL"
        # Check if list contains dicts (JSONB) or strings (TEXT[])
        if is_jsonb or (val and isinstance(val[0], dict)):
            # JSONB array format
            json_str = json.dumps(val).replace("'", "''")
            return f"'{json_str}'"
        else:
            # TEXT[] array format: ARRAY['item1', 'item2']
            escaped = [f"'{str(v).replace(chr(39), chr(39)+chr(39))}'" for v in val if v]
            return f"ARRAY[{', '.join(escaped)}]"
    if isinstance(val, dict):
        # JSONB format
        return f"'{json.dumps(val).replace(chr(39), chr(39)+chr(39))}'"
    # String - escape single quotes
    val = str(val).replace("'", "''")
    return f"'{val}'"

def parse_hotel_sheet(sheet, sheet_name):
    """Parse a single hotel sheet and return a dictionary of data"""
    hotel = {
        'excel_sheet_name': sheet_name,
    }

    # A6: Hotel name
    hotel['hotel_name'] = clean_value(sheet.cell(row=6, column=1).value)

    # A9: Owner LLC
    hotel['owner_llc'] = clean_value(sheet.cell(row=9, column=1).value)

    # A10: Operating Company LLC
    hotel['operating_company_llc'] = clean_value(sheet.cell(row=10, column=1).value)

    # A12: Address street
    hotel['address_street'] = clean_value(sheet.cell(row=12, column=1).value)

    # A13: City, State ZIP
    city_state_zip = clean_value(sheet.cell(row=13, column=1).value)
    city, state, zipcode = parse_city_state_zip(city_state_zip)
    hotel['address_city'] = city
    hotel['address_state'] = state
    hotel['address_zip'] = zipcode

    # A14: Phone
    phone = clean_value(sheet.cell(row=14, column=1).value)
    if phone:
        # Clean phone - just get digits and format
        phone = phone.replace('(', '').replace(')', '').replace('-', '').replace(' ', '')
        hotel['phone'] = phone
    else:
        hotel['phone'] = None

    # A15: Website
    hotel['website'] = clean_value(sheet.cell(row=15, column=1).value)

    # A17: County/Submarket
    county_submarket = clean_value(sheet.cell(row=17, column=1).value)
    county, submarket = parse_county_submarket(county_submarket)
    hotel['county'] = county
    hotel['submarket'] = submarket

    # A19: Year Built
    hotel['year_built'] = extract_year(clean_value(sheet.cell(row=19, column=1).value))

    # A21: Parking
    parking_text = clean_value(sheet.cell(row=21, column=1).value)
    if parking_text:
        hotel['parking_spaces'] = extract_number(parking_text)
        if 'surface' in parking_text.lower():
            hotel['parking_type'] = 'Surface'
        elif 'garage' in parking_text.lower():
            hotel['parking_type'] = 'Garage'
        elif 'valet' in parking_text.lower():
            hotel['parking_type'] = 'Valet'
        else:
            hotel['parking_type'] = parking_text
    else:
        hotel['parking_spaces'] = None
        hotel['parking_type'] = None

    # A23: Number of buildings
    buildings_text = clean_value(sheet.cell(row=23, column=1).value)
    hotel['num_buildings'] = extract_number(buildings_text)

    # A24: Number of stories/floors
    stories_text = clean_value(sheet.cell(row=24, column=1).value)
    hotel['num_stories'] = extract_number(stories_text)

    # A25: Total square footage
    sqft_text = clean_value(sheet.cell(row=25, column=1).value)
    if sqft_text:
        # Remove commas and extract number
        sqft_text = sqft_text.replace(',', '')
        hotel['total_sq_ft'] = extract_number(sqft_text)
    else:
        hotel['total_sq_ft'] = None

    # A27: Acreage
    acreage_text = clean_value(sheet.cell(row=27, column=1).value)
    if acreage_text:
        # Match decimal numbers like 1.98, 0.378, etc.
        match = re.search(r'(\d+\.?\d*)', str(acreage_text))
        if match:
            try:
                hotel['acreage'] = float(match.group(1))
            except ValueError:
                hotel['acreage'] = None
        else:
            hotel['acreage'] = None
    else:
        hotel['acreage'] = None

    # A29: Purchase Date
    purchase_text = clean_value(sheet.cell(row=29, column=1).value)
    if purchase_text:
        # Extract the date part after the colon
        if ':' in purchase_text:
            hotel['purchase_date'] = purchase_text.split(':', 1)[1].strip() or None
        else:
            hotel['purchase_date'] = None
    else:
        hotel['purchase_date'] = None

    # A30: Open Date
    open_text = clean_value(sheet.cell(row=30, column=1).value)
    if open_text:
        if ':' in open_text:
            hotel['open_date'] = open_text.split(':', 1)[1].strip() or None
        else:
            hotel['open_date'] = None
    else:
        hotel['open_date'] = None

    # A32: Marsha Code
    marsha_text = clean_value(sheet.cell(row=32, column=1).value)
    if marsha_text and ':' in marsha_text:
        hotel['marsha_code'] = marsha_text.split(':', 1)[1].strip() or None
    else:
        hotel['marsha_code'] = None

    # A33: STR Code
    str_text = clean_value(sheet.cell(row=33, column=1).value)
    if str_text and ':' in str_text:
        hotel['str_code'] = str_text.split(':', 1)[1].strip() or None
    else:
        hotel['str_code'] = None

    # A34: FEIN
    fein_text = clean_value(sheet.cell(row=34, column=1).value)
    if fein_text and ':' in fein_text:
        hotel['fein'] = fein_text.split(':', 1)[1].strip() or None
    else:
        hotel['fein'] = None

    # Room Mix (Rows 39-44 typically)
    room_mix = []
    total_rooms = 0
    for row in range(39, 50):
        room_type = clean_value(sheet.cell(row=row, column=1).value)
        room_count = sheet.cell(row=row, column=2).value
        room_sqft = sheet.cell(row=row, column=4).value

        if room_type and room_count and room_type.lower() not in ['total', 'total/weighted avg', 'meeting room', 'meeting room(s)']:
            if isinstance(room_count, (int, float)) and room_count > 0:
                room_data = {
                    'type': room_type,
                    'count': int(room_count)
                }
                if isinstance(room_sqft, (int, float)):
                    room_data['sq_ft'] = int(room_sqft)
                room_mix.append(room_data)
                total_rooms += int(room_count)

        # Check for total row
        if room_type and 'total' in room_type.lower():
            # Try to get the actual total from cell B
            if isinstance(room_count, (int, float)):
                total_rooms = int(room_count)
            break

    hotel['room_mix'] = room_mix if room_mix else None
    hotel['total_rooms'] = total_rooms if total_rooms > 0 else None

    # Meeting Rooms (starting around row 46-47)
    meeting_rooms = []
    for row in range(46, 55):
        room_name = clean_value(sheet.cell(row=row, column=1).value)
        room_sqft = sheet.cell(row=row, column=2).value
        room_dims = clean_value(sheet.cell(row=row, column=3).value)

        if room_name and room_name.lower() not in ['meeting room(s)', 'features', 'sq footage', ''] and isinstance(room_sqft, (int, float)):
            meeting_rooms.append({
                'name': room_name,
                'sq_ft': int(room_sqft),
                'dimensions': room_dims
            })

    hotel['meeting_rooms'] = meeting_rooms if meeting_rooms else None

    # Features/Amenities (Rows 51-54)
    features = []
    for row in range(50, 56):
        feature = clean_value(sheet.cell(row=row, column=1).value)
        if feature and feature.startswith('ª'):
            features.append(feature[1:].strip())
    hotel['features_amenities'] = '\n'.join(features) if features else None

    # Renovations (around row 59-62)
    renovations_parts = []
    for row in range(58, 63):
        text = clean_value(sheet.cell(row=row, column=1).value)
        desc = clean_value(sheet.cell(row=row, column=4).value)
        if text and text not in ['Renovations', 'Item']:
            if desc:
                renovations_parts.append(f"{text}: {desc}")
            elif ':' in text:
                renovations_parts.append(text)
    hotel['renovations'] = '\n'.join(renovations_parts) if renovations_parts else None

    # Franchise Brand and Fees (rows 64-66)
    hotel['franchise_brand'] = clean_value(sheet.cell(row=65, column=1).value)
    hotel['franchise_fees'] = clean_value(sheet.cell(row=65, column=4).value)

    # Common Area/Retail (rows 68-70)
    common_parts = []
    for row in range(68, 75):
        text = clean_value(sheet.cell(row=row, column=1).value)
        desc = clean_value(sheet.cell(row=row, column=4).value)
        if text and text not in ['Common Area/Association/Information/Retail', 'Common Area/Association/Retail', 'Item', 'Loan Information']:
            if desc:
                common_parts.append(f"{text}: {desc}")
            else:
                common_parts.append(text)
    hotel['common_area_retail'] = '\n'.join(common_parts) if common_parts else None

    # Lender Info (rows 71-75)
    lender_parts = []
    for row in range(70, 80):
        text = clean_value(sheet.cell(row=row, column=1).value)
        desc = clean_value(sheet.cell(row=row, column=4).value)
        if text and text.startswith('Lender'):
            if desc:
                lender_parts.append(f"{text} {desc}")
            else:
                lender_parts.append(text)
    hotel['lender_info'] = '\n'.join(lender_parts) if lender_parts else None

    # Competitive Set (Column F, rows 11-20)
    comp_set = []
    for row in range(11, 25):
        comp = clean_value(sheet.cell(row=row, column=6).value)
        if comp and comp not in ['Hotel', 'Competitive Set Information']:
            comp_set.append(comp)
    hotel['competitive_set'] = comp_set if comp_set else None

    # Management Company (Column F around row 25-26)
    for row in range(25, 30):
        text = clean_value(sheet.cell(row=row, column=6).value)
        if text and text not in ['Management Company']:
            hotel['management_company'] = text
            break
    else:
        hotel['management_company'] = None

    return hotel

def should_skip_sheet(sheet_name):
    """Check if a sheet should be skipped"""
    skip_patterns = ['template', 'SOLD HOTELS']
    name_lower = sheet_name.lower()
    for pattern in skip_patterns:
        if pattern.lower() in name_lower:
            return True
    return False

def generate_insert_sql(hotel):
    """Generate an INSERT statement for a hotel"""
    columns = [
        'hotel_name', 'owner_llc', 'operating_company_llc',
        'address_street', 'address_city', 'address_state', 'address_zip',
        'phone', 'website', 'county', 'submarket',
        'year_built', 'parking_spaces', 'parking_type',
        'num_buildings', 'num_stories', 'total_sq_ft', 'acreage',
        'purchase_date', 'open_date',
        'marsha_code', 'str_code', 'fein',
        'total_rooms', 'room_mix', 'meeting_rooms',
        'features_amenities', 'renovations',
        'franchise_brand', 'franchise_fees',
        'common_area_retail', 'lender_info',
        'competitive_set', 'management_company',
        'excel_sheet_name'
    ]

    values = []
    for col in columns:
        val = hotel.get(col)
        if col in ['room_mix', 'meeting_rooms']:
            # JSONB columns
            values.append(escape_sql(val) if val else "NULL")
        elif col == 'competitive_set':
            # Array column
            values.append(escape_sql(val) if val else "NULL")
        else:
            values.append(escape_sql(val))

    return f"INSERT INTO public.hotel_fact_sheets ({', '.join(columns)})\nVALUES ({', '.join(values)});"

def main():
    print(f"Loading Excel file: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH)

    hotels = []
    skipped = []

    for sheet_name in wb.sheetnames:
        if should_skip_sheet(sheet_name):
            skipped.append(sheet_name)
            continue

        sheet = wb[sheet_name]

        # Check if sheet has a hotel name in A6
        hotel_name = clean_value(sheet.cell(row=6, column=1).value)
        if not hotel_name:
            skipped.append(f"{sheet_name} (no hotel name)")
            continue

        print(f"Parsing: {sheet_name} -> {hotel_name}")
        hotel = parse_hotel_sheet(sheet, sheet_name)
        hotels.append(hotel)

    print(f"\nParsed {len(hotels)} hotels")
    print(f"Skipped {len(skipped)} sheets: {skipped}")

    # Generate SQL
    sql_statements = [
        "-- Migration: 029_import_hotel_fact_sheets.sql",
        "-- Auto-generated import of hotel fact sheet data",
        "",
        "-- Clear existing data (optional - comment out if you want to keep existing data)",
        "-- TRUNCATE TABLE public.hotel_fact_sheets;",
        ""
    ]

    for hotel in hotels:
        sql_statements.append(generate_insert_sql(hotel))
        sql_statements.append("")

    # Write SQL file
    sql_content = '\n'.join(sql_statements)

    with open(OUTPUT_SQL_PATH, 'w') as f:
        f.write(sql_content)

    print(f"\nGenerated SQL file: {OUTPUT_SQL_PATH}")
    print(f"Total INSERT statements: {len(hotels)}")

if __name__ == "__main__":
    main()
